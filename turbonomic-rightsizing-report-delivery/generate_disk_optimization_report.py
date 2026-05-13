#!/usr/bin/env python3
"""
Turbonomic Disk Optimization Report Generator v2 (Consolidated Excel Format)

Generates consolidated disk tier optimization reports with:
- Single Excel workbook with multiple sheets (Summary, Comprehensive, Dev, UAT, Pre-Prod, Prod, DR, Unmapped)
- Professional formatting matching Monthly Action Plan style
- Clickable action links with corrected URL format
- Policy enforcement: DEV/UAT Premium SSD = violation, Pre-Prod/Prod/DR = review required

Requirements:
    pip install requests pandas openpyxl

Usage:
    python3 generate_disk_optimization_report_v2.py --url https://your-turbo-instance.com --jsessionid <SESSION_ID> --output-dir reports/
"""

import requests
import json
import argparse
import sys
import os
from datetime import datetime
from typing import List, Dict, Optional
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

try:
    import pandas as pd
    from openpyxl import Workbook
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False


class TurbonomicDiskOptimizationReport:
    """Generate consolidated disk optimization reports from Turbonomic API with policy enforcement."""
    
    def __init__(self, turbo_url: str, jsessionid: str, customer_mapping_file: Optional[str] = None):
        self.turbo_url = turbo_url.rstrip('/')
        self.session = requests.Session()
        self.session.cookies.set('JSESSIONID', jsessionid)
        self.session.headers.update({
            'Content-Type': 'application/json',
            'Accept': 'application/json'
        })
        self.customer_mapping = self._load_customer_mapping(customer_mapping_file)
        
        # Environment sort order for consistent ordering
        self.env_sort_order = {'Dev': 1, 'UAT': 2, 'Pre-Prod': 3, 'Prod': 4, 'DR': 5, 'Unmapped': 6}
    
    def _load_customer_mapping(self, mapping_file: Optional[str]) -> Dict[str, str]:
        """Load customer ID to name mapping from JSON file."""
        if not mapping_file:
            return {}
        
        try:
            with open(mapping_file, 'r', encoding='utf-8') as f:
                mapping = json.load(f)
                print(f"✓ Loaded {len(mapping)} customer mappings from {mapping_file}")
                return mapping
        except FileNotFoundError:
            print(f"Warning: Customer mapping file not found: {mapping_file}")
            return {}
        except json.JSONDecodeError as e:
            print(f"Warning: Invalid JSON in customer mapping file: {e}")
            return {}
        except Exception as e:
            print(f"Warning: Error loading customer mapping file: {e}")
            return {}
    
    def _map_customer_name(self, customer_id: str) -> str:
        """Map customer ID to friendly name using the mapping file."""
        if not customer_id or customer_id == 'N/A':
            return 'N/A'
        
        if customer_id in self.customer_mapping:
            return self.customer_mapping[customer_id]
        
        return customer_id
    
    def get_storage_actions(self, fetch_all: bool = True) -> List[Dict]:
        """Fetch storage-related actions with pagination support."""
        url = f"{self.turbo_url}/api/v3/markets/Market/actions"
        
        # Query specifically for storage tier actions
        payload = {
            "actionStateList": ["READY", "ACCEPTED", "QUEUED", "IN_PROGRESS"],
            "actionTypeList": ["RESIZE", "SCALE", "RECONFIGURE"],
            "relatedEntityTypes": ["VirtualVolume", "Storage", "Volume", "VirtualMachineVolume"],
            "environmentType": "CLOUD",
            "detailLevel": "EXECUTION"
        }
        
        all_actions = []
        cursor = 0
        page_size = 500
        
        print(f"Fetching storage actions from {url}...")
        
        try:
            while True:
                params = {
                    "ascending": "false",
                    "cursor": str(cursor),
                    "disable_hateoas": "true",
                    "forceExpansionOfAggregatedEntities": "true",
                    "order_by": "savings",
                    "limit": str(page_size)
                }
                
                response = self.session.post(url, json=payload, params=params)
                response.raise_for_status()
                
                data = response.json()
                actions = data if isinstance(data, list) else []
                
                if not actions:
                    break
                
                all_actions.extend(actions)
                print(f"  Retrieved {len(actions)} actions (total: {len(all_actions)})")
                
                if len(actions) < page_size or not fetch_all:
                    break
                
                cursor += page_size
            
            print(f"Total storage actions retrieved: {len(all_actions)}")
            return all_actions
            
        except requests.exceptions.RequestException as e:
            print(f"Error fetching storage actions: {e}")
            if hasattr(e, 'response') and e.response is not None:
                print(f"Response: {e.response.text}")
            return []
    
    def _get_cloud_provider(self, action: Dict) -> str:
        """Determine cloud provider from action data."""
        target = action.get('target', {})
        discovered_by = target.get('discoveredBy', {})
        probe_type = discovered_by.get('type', '').lower()
        
        if 'azure' in probe_type:
            return 'Azure'
        elif 'aws' in probe_type:
            return 'AWS'
        elif 'gcp' in probe_type or 'google' in probe_type:
            return 'GCP'
        
        aspects = target.get('aspects', {})
        if 'cloudAspect' in aspects:
            return 'Azure'
        
        return 'Unknown'
    
    def _get_business_account(self, action: Dict) -> str:
        """Extract BusinessAccount name from cloudAspect."""
        target = action.get('target', {})
        aspects = target.get('aspects', {})
        cloud_aspect = aspects.get('cloudAspect', {})
        business_account = cloud_aspect.get('businessAccount', {})
        
        return business_account.get('displayName', 'N/A')
    
    def _parse_environment_from_account(self, account_name: str) -> Optional[str]:
        """Parse environment from BusinessAccount name."""
        account_upper = account_name.upper()
        
        if any(x in account_upper for x in ['PRODUCTION', ' PRD ', '-PRD-', '_PRD_', 'PRD-', '-PRD']):
            return 'Prod'
        
        if any(x in account_upper for x in ['PRE-PROD', 'PREPROD', 'PRE-PRODUCTION', 'PREPRODUCTION']):
            return 'Pre-Prod'
        
        if any(x in account_upper for x in [' DR ', '-DR-', '_DR_', 'DR-', '-DR', 'DISASTER RECOVERY', 'DISASTERRECOVERY']):
            return 'DR'
        
        if any(x in account_upper for x in ['UAT', 'USER ACCEPTANCE', 'USERACCEPTANCE']):
            return 'UAT'
        
        if any(x in account_upper for x in ['DEVELOPMENT', ' DEV ', '-DEV-', '_DEV_', 'DEV-', '-DEV']):
            return 'Dev'
        
        return None
    
    def _determine_environment(self, action: Dict) -> str:
        """Determine environment with fallback logic."""
        # Try BusinessAccount first
        account_name = self._get_business_account(action)
        if account_name != 'N/A':
            env_from_account = self._parse_environment_from_account(account_name)
            if env_from_account:
                return env_from_account
        
        # Try to get VM and check its tags
        vm_name = self._get_attached_vm(action)
        if vm_name != 'N/A':
            env_from_vm = self._parse_environment_from_account(vm_name)
            if env_from_vm:
                return env_from_vm
        
        # Fall back to environment tag
        target = action.get('target', {})
        env_tag = self._get_tag_value(target, 'environment')
        
        if env_tag and env_tag != 'N/A':
            env_upper = env_tag.upper()
            if 'PROD' in env_upper and 'PRE' not in env_upper:
                return 'Prod'
            elif 'PRE-PROD' in env_upper or 'PREPROD' in env_upper:
                return 'Pre-Prod'
            elif 'UAT' in env_upper:
                return 'UAT'
            elif 'DEV' in env_upper:
                return 'Dev'
            elif 'DR' in env_upper:
                return 'DR'
            # If tag exists but doesn't match patterns, categorize as Unmapped
        
        return 'Unmapped'
    
    def _get_tag_value(self, entity: Dict, tag_key: str) -> str:
        """Extract tag value from entity."""
        tags = entity.get('tags', {})
        
        if isinstance(tags, dict):
            for key, value in tags.items():
                if key.lower() == tag_key.lower():
                    if isinstance(value, list) and value:
                        return value[0] if value[0] else 'N/A'
                    return str(value) if value else 'N/A'
        
        return 'N/A'
    
    def _extract_disk_tier(self, tier_string: str) -> str:
        """Extract disk tier from string."""
        if not tier_string or tier_string == 'N/A':
            return 'N/A'
        
        tier_upper = tier_string.upper()
        
        if 'PREMIUM' in tier_upper and 'SSD' in tier_upper:
            return 'Premium SSD'
        elif 'STANDARD' in tier_upper and 'SSD' in tier_upper:
            return 'Standard SSD'
        elif 'STANDARD' in tier_upper and 'HDD' in tier_upper:
            return 'Standard HDD'
        elif 'ULTRA' in tier_upper:
            return 'Ultra SSD'
        
        return tier_string
    
    def _get_attached_vm(self, action: Dict) -> str:
        """Get the VM name that the disk is attached to."""
        # Method 1: Check virtualDisks array (most reliable for disk actions)
        virtual_disks = action.get('virtualDisks', [])
        if virtual_disks and isinstance(virtual_disks, list):
            for vdisk in virtual_disks:
                if isinstance(vdisk, dict):
                    attached_vm = vdisk.get('attachedVirtualMachine', {})
                    if attached_vm and isinstance(attached_vm, dict):
                        vm_name = attached_vm.get('displayName', '')
                        if vm_name:
                            return vm_name
        
        # Method 2: Try to get from target providers array
        target = action.get('target', {})
        providers = target.get('providers', [])
        if providers and isinstance(providers, list):
            for provider in providers:
                if isinstance(provider, dict):
                    vm_name = provider.get('displayName', '')
                    if vm_name:
                        return vm_name
        
        # Method 3: Try to get from target consumers array
        consumers = target.get('consumers', [])
        if consumers and isinstance(consumers, list):
            for consumer in consumers:
                if isinstance(consumer, dict):
                    entity_type = consumer.get('className', '')
                    if 'VirtualMachine' in entity_type or 'VM' in entity_type:
                        vm_name = consumer.get('displayName', '')
                        if vm_name:
                            return vm_name
        
        # Method 4: Parse from action details text as last resort
        details = action.get('details', '')
        if details and 'attached to' in details.lower():
            import re
            match = re.search(r'attached to\s+([^\s,\.]+)', details, re.IGNORECASE)
            if match:
                return match.group(1)
        
        return 'N/A'
    
    def _calculate_monthly_savings(self, action: Dict) -> float:
        """Calculate monthly savings from action."""
        stats = action.get('stats', [])
        for stat in stats:
            if isinstance(stat, dict):
                stat_name = stat.get('name', '').lower()
                value = stat.get('value', 0)
                
                # In Turbonomic API: Positive cost value = SAVINGS
                if 'cost' in stat_name and value > 0:  # Positive = savings
                    hourly_rate = abs(float(value))
                    return hourly_rate * 730
        
        return 0.0
    
    def _get_action_link(self, action: Dict) -> str:
        """Generate link to entity page with actions tab in Turbonomic UI."""
        target = action.get('target', {})
        entity_uuid = target.get('uuid', '')
        if entity_uuid:
            return f"{self.turbo_url}/app/#/view/main/live/{entity_uuid}/overview?selectedTab=actions"
        return 'N/A'
    
    def _determine_policy_status(self, environment: str, current_tier: str, recommended_tier: str) -> str:
        """Determine if action is a policy violation or requires review."""
        if environment in ['Dev', 'UAT'] and 'Premium' in current_tier and 'Standard' in recommended_tier:
            return 'POLICY VIOLATION'
        elif environment in ['Pre-Prod', 'Prod', 'DR']:
            return 'REVIEW REQUIRED'
        else:
            return 'RECOMMENDED'
    
    def generate_report_data(self, azure_only: bool = True) -> List[Dict]:
        """Generate report data from Turbonomic storage actions."""
        actions = self.get_storage_actions()
        
        report_data = []
        
        for action in actions:
            target = action.get('target', {})
            
            # Filter for Azure only
            cloud_provider = self._get_cloud_provider(action)
            if azure_only and cloud_provider != 'Azure':
                continue
            
            # Get disk information
            disk_name = target.get('displayName', 'N/A')
            vm_name = self._get_attached_vm(action)
            
            # Determine environment
            environment = self._determine_environment(action)
            
            # Get tier information
            current_entity = action.get('currentEntity', {})
            new_entity = action.get('newEntity', {})
            current_tier = self._extract_disk_tier(current_entity.get('displayName', 'N/A'))
            recommended_tier = self._extract_disk_tier(new_entity.get('displayName', 'N/A'))
            
            # Skip if not a tier change
            if current_tier == recommended_tier or current_tier == 'N/A' or recommended_tier == 'N/A':
                continue
            
            # Get customer ID and map to friendly name
            customer_id = self._get_tag_value(target, 'CustomerID')
            customer_name = self._map_customer_name(customer_id)
            
            # Calculate savings
            monthly_savings = self._calculate_monthly_savings(action)
            
            # Determine policy status
            policy_status = self._determine_policy_status(environment, current_tier, recommended_tier)
            
            # Build justification
            if policy_status == 'POLICY VIOLATION':
                justification = 'POLICY VIOLATION: Premium SSD not allowed in DEV/UAT environments'
            elif policy_status == 'REVIEW REQUIRED':
                justification = f"Review Required: {action.get('details', 'Disk tier optimization')}"
            else:
                justification = action.get('details', 'Disk tier optimization recommended')
            
            # Build report row
            row = {
                'Disk Name': disk_name,
                'Attached VM': vm_name,
                'Customer ID': customer_id,
                'Customer Friendly Name': customer_name,
                'Current Tier': current_tier,
                'Recommended Tier': recommended_tier,
                'Monthly Savings': monthly_savings,
                'Environment': environment,
                'Policy Status': policy_status,
                'Justification': justification,
                'Action Details Link': self._get_action_link(action),
                'Business Account': self._get_business_account(action),
                'Cloud Provider': cloud_provider,
                'Action State': action.get('actionState', 'N/A'),
                'Risk': action.get('risk', {}).get('severity', 'N/A'),
                'UUID': target.get('uuid', 'N/A')
            }
            
            report_data.append(row)
        
        return report_data
    
    def _sort_data(self, data: List[Dict]) -> List[Dict]:
        """Sort data by environment then by savings (largest to smallest)."""
        return sorted(data, key=lambda x: (
            self.env_sort_order.get(x.get('Environment', 'Unmapped'), 999),
            -x.get('Monthly Savings', 0)
        ))
    
    def export_consolidated_excel(self, all_data: List[Dict], filename: str):
        """Export consolidated report to Excel with multiple sheets and professional formatting."""
        if not PANDAS_AVAILABLE:
            print("Error: pandas and openpyxl required for Excel export")
            return
        
        if not all_data:
            print(f"No data to export to {filename}")
            return
        
        # Sort all data
        all_data = self._sort_data(all_data)
        
        # Create Excel writer
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            # Create Summary sheet
            self._create_summary_sheet(writer, all_data)
            
            # Create Comprehensive sheet (all data)
            df_comprehensive = pd.DataFrame(all_data)
            df_comprehensive.to_excel(writer, sheet_name='Comprehensive', index=False)
            self._format_data_sheet(writer.book['Comprehensive'])
            
            # Create environment-specific sheets
            environments = ['Dev', 'UAT', 'Pre-Prod', 'Prod', 'DR', 'Unmapped']
            for env in environments:
                env_data = [row for row in all_data if row['Environment'] == env]
                if env_data:
                    df_env = pd.DataFrame(env_data)
                    df_env.to_excel(writer, sheet_name=env, index=False)
                    self._format_data_sheet(writer.book[env])
        
        print(f"\n✓ Consolidated disk optimization report exported to: {filename}")
        print(f"  Sheets: Summary, Comprehensive, {', '.join(environments)}")
    
    def _create_summary_sheet(self, writer, all_data: List[Dict]):
        """Create summary sheet with statistics."""
        # Calculate overall statistics
        total_actions = len(all_data)
        policy_violations = len([r for r in all_data if r['Policy Status'] == 'POLICY VIOLATION'])
        review_required = len([r for r in all_data if r['Policy Status'] == 'REVIEW REQUIRED'])
        recommended = len([r for r in all_data if r['Policy Status'] == 'RECOMMENDED'])
        
        total_monthly_savings = sum(r['Monthly Savings'] for r in all_data)
        
        # Environment breakdown
        env_breakdown = {}
        env_savings = {}
        env_violations = {}
        env_reviews = {}
        
        for row in all_data:
            env = row.get('Environment', 'Unknown')
            env_breakdown[env] = env_breakdown.get(env, 0) + 1
            env_savings[env] = env_savings.get(env, 0) + row['Monthly Savings']
            
            if row['Policy Status'] == 'POLICY VIOLATION':
                env_violations[env] = env_violations.get(env, 0) + 1
            elif row['Policy Status'] == 'REVIEW REQUIRED':
                env_reviews[env] = env_reviews.get(env, 0) + 1
        
        # Build summary data
        summary_rows = []
        
        # Overall summary
        summary_rows.append({
            'Category': 'OVERALL SUMMARY',
            'Count': '',
            'Policy Violations': '',
            'Review Required': '',
            'Recommended': '',
            'Monthly Savings': ''
        })
        summary_rows.append({
            'Category': 'Total Disk Optimizations',
            'Count': total_actions,
            'Policy Violations': policy_violations,
            'Review Required': review_required,
            'Recommended': recommended,
            'Monthly Savings': f"${total_monthly_savings:,.2f}"
        })
        summary_rows.append({
            'Category': '',
            'Count': '',
            'Policy Violations': '',
            'Review Required': '',
            'Recommended': '',
            'Monthly Savings': ''
        })
        
        # Environment breakdown
        summary_rows.append({
            'Category': 'ENVIRONMENT BREAKDOWN',
            'Count': '',
            'Policy Violations': '',
            'Review Required': '',
            'Recommended': '',
            'Monthly Savings': ''
        })
        
        for env in sorted(env_breakdown.keys(), key=lambda x: self.env_sort_order.get(x, 999)):
            count = env_breakdown[env]
            savings = env_savings.get(env, 0)
            violations = env_violations.get(env, 0)
            reviews = env_reviews.get(env, 0)
            recommended_count = count - violations - reviews
            
            summary_rows.append({
                'Category': env,
                'Count': count,
                'Policy Violations': violations,
                'Review Required': reviews,
                'Recommended': recommended_count,
                'Monthly Savings': f"${savings:,.2f}"
            })
        
        df_summary = pd.DataFrame(summary_rows)
        df_summary.to_excel(writer, sheet_name='Summary', index=False)
        
        # Format summary sheet
        ws = writer.book['Summary']
        self._format_summary_sheet(ws)
    
    def _format_summary_sheet(self, ws):
        """Format the summary sheet."""
        # Header formatting
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Section header formatting
        section_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        section_font = Font(bold=True, size=11)
        
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            if row[0].value in ['OVERALL SUMMARY', 'ENVIRONMENT BREAKDOWN']:
                for cell in row:
                    cell.fill = section_fill
                    cell.font = section_font
        
        # Environment row formatting
        env_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            if row[0].value in ['Dev', 'UAT', 'Pre-Prod', 'Prod', 'DR', 'Unmapped']:
                for cell in row:
                    cell.fill = env_fill
        
        # Column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 20
        
        # Borders
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=6):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(vertical='center', wrap_text=True)
        
        # Freeze panes
        ws.freeze_panes = 'A2'
    
    def _format_data_sheet(self, ws):
        """Format data sheets (Comprehensive and environment-specific)."""
        # Header formatting
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Column widths
        column_widths = {
            'A': 35,  # Disk Name
            'B': 35,  # Attached VM
            'C': 20,  # Customer ID
            'D': 30,  # Customer Friendly Name
            'E': 20,  # Current Tier
            'F': 20,  # Recommended Tier
            'G': 18,  # Monthly Savings
            'H': 15,  # Environment
            'I': 20,  # Policy Status
            'J': 50,  # Justification
            'K': 50,  # Action Details Link
            'L': 30,  # Business Account
            'M': 15,  # Cloud Provider
            'N': 15,  # Action State
            'O': 12,  # Risk
            'P': 38   # UUID
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # Borders and alignment
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=16):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(vertical='center', wrap_text=True)
        
        # Make Action Details Link column clickable (column K)
        for row in range(2, ws.max_row + 1):
            link_cell = ws[f'K{row}']
            if link_cell.value and link_cell.value != 'N/A':
                link_cell.hyperlink = link_cell.value
                link_cell.style = 'Hyperlink'
        
        # Highlight policy violations
        violation_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        for row in range(2, ws.max_row + 1):
            policy_status = ws[f'I{row}'].value
            if policy_status and 'POLICY VIOLATION' in str(policy_status):
                for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P']:
                    ws[f'{col}{row}'].fill = violation_fill
        
        # Freeze panes
        ws.freeze_panes = 'A2'
    
    def print_summary(self, report_data: List[Dict], title: str = "DISK OPTIMIZATION REPORT SUMMARY"):
        """Print report summary statistics."""
        if not report_data:
            print(f"No data in {title}")
            return
        
        total_actions = len(report_data)
        policy_violations = len([r for r in report_data if r['Policy Status'] == 'POLICY VIOLATION'])
        review_required = len([r for r in report_data if r['Policy Status'] == 'REVIEW REQUIRED'])
        recommended = len([r for r in report_data if r['Policy Status'] == 'RECOMMENDED'])
        
        total_monthly_savings = sum(r['Monthly Savings'] for r in report_data)
        
        print("\n" + "="*80)
        print(title)
        print("="*80)
        print(f"Total Disk Optimizations: {total_actions}")
        print(f"  - Policy Violations: {policy_violations}")
        print(f"  - Review Required: {review_required}")
        print(f"  - Recommended: {recommended}")
        print(f"\nEstimated Monthly Savings: ${total_monthly_savings:,.2f}")
        
        environments = {}
        for row in report_data:
            env = row.get('Environment', 'Unknown')
            environments[env] = environments.get(env, 0) + 1
        
        if environments:
            print(f"\nEnvironment Breakdown:")
            for env in sorted(environments.keys(), key=lambda x: self.env_sort_order.get(x, 999)):
                print(f"  - {env}: {environments[env]}")
        
        print("="*80)


def main():
    parser = argparse.ArgumentParser(
        description='Generate Turbonomic disk optimization reports v2 (Consolidated Excel Format)',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Generate consolidated disk optimization report
  python3 generate_disk_optimization_report_v2.py --url https://turbo.example.com --jsessionid <ID> --output-dir reports/
  
  # With customer mapping
  python3 generate_disk_optimization_report_v2.py --url https://turbo.example.com --jsessionid <ID> --customer-mapping customer_mapping.json
  
  # Custom output filename
  python3 generate_disk_optimization_report_v2.py --url https://turbo.example.com --jsessionid <ID> --output disk_opt_jan2026.xlsx
        """
    )
    
    parser.add_argument('--url', required=True, help='Turbonomic instance URL')
    parser.add_argument('--jsessionid', required=True, help='JSESSIONID from login')
    parser.add_argument('--output-dir', default='.', help='Output directory for reports (default: current directory)')
    parser.add_argument('--output', help='Custom output filename (default: Disk_Optimization_Report_TIMESTAMP.xlsx)')
    parser.add_argument('--all-clouds', action='store_true', help='Include all cloud providers (not just Azure)')
    parser.add_argument('--customer-mapping', default='customer_mapping.json', help='Path to customer ID to name mapping JSON file (default: customer_mapping.json)')
    
    args = parser.parse_args()
    
    try:
        # Validate and create output directory
        output_dir = args.output_dir
        
        if os.path.isfile(output_dir):
            print(f"Error: '{output_dir}' is a file, not a directory.")
            print(f"Use --output-dir to specify a directory, not a filename.")
            print(f"Example: --output-dir ./reports/")
            return 1
        
        # Create output directory
        os.makedirs(output_dir, exist_ok=True)
        
        # Use customer mapping if file exists
        customer_mapping_file = args.customer_mapping if os.path.exists(args.customer_mapping) else None
        
        report = TurbonomicDiskOptimizationReport(args.url, args.jsessionid, customer_mapping_file)
        
        print("Generating consolidated disk optimization report...")
        all_data = report.generate_report_data(azure_only=not args.all_clouds)
        
        if not all_data:
            print("No disk optimization actions found. Exiting.")
            return 1
        
        # Generate consolidated report
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        if args.output:
            output_file = os.path.join(output_dir, args.output)
        else:
            output_file = os.path.join(output_dir, f"Disk_Optimization_Report_{timestamp}.xlsx")
        
        report.export_consolidated_excel(all_data, output_file)
        report.print_summary(all_data, "CONSOLIDATED DISK OPTIMIZATION REPORT SUMMARY")
        
        print(f"\n{'='*80}")
        print("Report generation complete!")
        print(f"{'='*80}")
        
        return 0
        
    except KeyboardInterrupt:
        print("\n\nOperation cancelled by user")
        return 1
    except Exception as e:
        print(f"\nError: {e}")
        import traceback
        traceback.print_exc()
        return 1


if __name__ == '__main__':
    sys.exit(main())

# Made with Bob
