#!/usr/bin/env python3
"""
Turbonomic Monthly Action Plan Report Generator

Generates a comprehensive monthly action plan with three categories:
1. Must-Do Actions: Policy violations (DEV/UAT Premium disks) + validated downsizing
2. Cost Optimization: Recommended downsizing across all environments
3. Reliability Investment: Prod upsizing with budget impact

Requirements:
    pip install requests pandas openpyxl

Usage:
    python3 generate_monthly_action_plan.py --url https://your-turbo-instance.com --jsessionid <SESSION_ID>
"""

import requests
import json
import argparse
import sys
import re
import os
from datetime import datetime
from typing import List, Dict, Optional, Tuple
from collections import defaultdict

try:
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False
    print("Warning: pandas and openpyxl not available. Install with: pip install pandas openpyxl")


class MonthlyActionPlanGenerator:
    """Generate monthly action plan report from Turbonomic API."""
    
    def __init__(self, turbo_url: str, jsessionid: str, customer_mapping_file: Optional[str] = None):
        self.turbo_url = turbo_url.rstrip('/')
        self.session = requests.Session()
        self.session.cookies.set('JSESSIONID', jsessionid)
        self.session.headers.update({
            'Content-Type': 'application/json',
            'Accept': 'application/json'
        })
        self.customer_mapping = self._load_customer_mapping(customer_mapping_file)
        
        # Environment sort order
        self.env_sort_order = {'Dev': 1, 'UAT': 2, 'Pre-Prod': 3, 'Prod': 4, 'DR': 5, 'Unmapped': 6}
        
        # Disable SSL warnings for self-signed certificates
        import urllib3
        urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
    
    def _load_customer_mapping(self, mapping_file: Optional[str]) -> Dict[str, str]:
        """Load customer ID to name mapping from JSON file."""
        if not mapping_file or not os.path.exists(mapping_file):
            return {}
        
        try:
            with open(mapping_file, 'r', encoding='utf-8') as f:
                mapping = json.load(f)
                print(f"✓ Loaded {len(mapping)} customer mappings")
                return mapping
        except Exception as e:
            print(f"Warning: Error loading customer mapping: {e}")
            return {}
    
    def _map_customer_name(self, customer_id: str) -> str:
        """Map customer ID to friendly name using the mapping file."""
        if not customer_id or customer_id == 'N/A':
            return 'N/A'
        
        # Try exact match
        if customer_id in self.customer_mapping:
            return self.customer_mapping[customer_id]
        
        # Return original ID if no mapping found
        return customer_id
    
    def get_vm_actions(self) -> List[Dict]:
        """Fetch VM rightsizing actions."""
        url = f"{self.turbo_url}/api/v3/markets/Market/actions"
        
        payload = {
            "actionStateList": ["READY", "QUEUED", "ACCEPTED"],
            "actionTypeList": ["RESIZE", "SCALE"],
            "relatedEntityTypes": ["VirtualMachine"],
            "environmentType": "CLOUD",
            "detailLevel": "EXECUTION"
        }
        
        all_actions = []
        cursor = 0
        page_size = 500
        
        print(f"Fetching VM rightsizing actions...")
        
        try:
            while True:
                params = {
                    "ascending": "false",
                    "cursor": str(cursor),
                    "limit": str(page_size)
                }
                
                response = self.session.post(url, json=payload, params=params, verify=False, timeout=30)
                response.raise_for_status()
                data = response.json()
                
                actions = data if isinstance(data, list) else []
                if not actions:
                    break
                
                all_actions.extend(actions)
                
                if len(actions) < page_size:
                    break
                
                cursor += page_size
            
            print(f"✓ Retrieved {len(all_actions)} VM actions")
            return all_actions
            
        except requests.exceptions.RequestException as e:
            print(f"Error fetching VM actions: {e}")
            return []
    
    def get_storage_actions(self) -> List[Dict]:
        """Fetch storage/disk optimization actions."""
        url = f"{self.turbo_url}/api/v3/markets/Market/actions"
        
        payload = {
            "actionStateList": ["READY", "ACCEPTED", "QUEUED"],
            "actionTypeList": ["RESIZE", "SCALE", "RECONFIGURE"],
            "relatedEntityTypes": ["VirtualVolume", "Storage", "Volume"],
            "environmentType": "CLOUD",
            "detailLevel": "EXECUTION"
        }
        
        all_actions = []
        cursor = 0
        page_size = 500
        
        print(f"Fetching storage optimization actions...")
        
        try:
            while True:
                params = {
                    "ascending": "false",
                    "cursor": str(cursor),
                    "limit": str(page_size)
                }
                
                response = self.session.post(url, json=payload, params=params, verify=False, timeout=30)
                response.raise_for_status()
                data = response.json()
                
                actions = data if isinstance(data, list) else []
                if not actions:
                    break
                
                all_actions.extend(actions)
                
                if len(actions) < page_size:
                    break
                
                cursor += page_size
            
            print(f"✓ Retrieved {len(all_actions)} storage actions")
            return all_actions
            
        except requests.exceptions.RequestException as e:
            print(f"Error fetching storage actions: {e}")
            return []
    
    def _determine_environment(self, action: Dict) -> str:
        """Determine environment from action."""
        target = action.get('target', {})
        
        # Try BusinessAccount first
        aspects = target.get('aspects', {})
        business_account = aspects.get('businessAccountAspect', {})
        account_name = business_account.get('businessAccount', {}).get('displayName', '')
        
        if account_name:
            env = self._parse_environment_from_name(account_name)
            if env:
                return env
        
        # Fall back to environment tag
        env_tag = self._get_tag_value(target, 'environment')
        if env_tag and env_tag != 'N/A':
            return self._normalize_environment(env_tag)
        
        return 'Unmapped'
    
    def _parse_environment_from_name(self, name: str) -> Optional[str]:
        """Parse environment from name string."""
        if not name:
            return None
        
        name_upper = name.upper()
        
        if 'PROD' in name_upper and 'PRE' not in name_upper:
            return 'Prod'
        elif 'PRE-PROD' in name_upper or 'PREPROD' in name_upper:
            return 'Pre-Prod'
        elif 'UAT' in name_upper:
            return 'UAT'
        elif 'DEV' in name_upper:
            return 'Dev'
        elif 'DR' in name_upper:
            return 'DR'
        
        return None
    
    def _normalize_environment(self, env: str) -> str:
        """Normalize environment string."""
        env_upper = env.upper()
        
        if 'PROD' in env_upper and 'PRE' not in env_upper:
            return 'Prod'
        elif 'PRE-PROD' in env_upper or 'PREPROD' in env_upper:
            return 'Pre-Prod'
        elif 'UAT' in env_upper:
            return 'UAT'
        elif 'DEV' in env_upper:
            return 'Dev'
        elif 'DR' in env_upper:
            return 'DR'
        
        # If environment doesn't match standard patterns, categorize as Unmapped
        return 'Unmapped'
    
    def _get_tag_value(self, entity: Dict, tag_key: str) -> str:
        """Extract tag value from entity."""
        tags = entity.get('tags', {})
        
        if isinstance(tags, dict):
            for key, value in tags.items():
                if key.lower() == tag_key.lower():
                    if isinstance(value, list) and value:
                        return value[0] if value[0] else 'N/A'
                    return str(value) if value else 'N/A'
        
        return 'N/A'
    
    def _determine_action_type(self, action: Dict) -> str:
        """Determine if action is upsize or downsize based on cost impact."""
        # First check cost stats - most reliable indicator
        # In Turbonomic API:
        # Positive cost value = SAVINGS (downsize - reducing resources saves money)
        # Negative cost value = ADDITIONAL COST (upsize - adding resources costs money)
        stats = action.get('stats', [])
        for stat in stats:
            if isinstance(stat, dict):
                stat_name = stat.get('name', '').lower()
                value = stat.get('value', 0)
                
                if 'cost' in stat_name and value != 0:
                    # Positive value = savings = Downsize
                    # Negative value = additional cost = Upsize
                    if value > 0:
                        return 'Downsize'
                    elif value < 0:
                        return 'Upsize'
        
        # Fallback to risk category
        risk = action.get('risk', {})
        risk_sub_category = risk.get('subCategory', '').lower()
        details = action.get('details', '').lower()
        
        if 'underutilized' in risk_sub_category or 'underutilized' in details:
            return 'Downsize'
        
        if 'overutilized' in risk_sub_category or 'overutilized' in details:
            return 'Upsize'
        
        return 'Resize'
    
    def _calculate_monthly_savings(self, action: Dict) -> float:
        """Calculate monthly savings from action (only for downsizing)."""
        stats = action.get('stats', [])
        for stat in stats:
            if isinstance(stat, dict):
                stat_name = stat.get('name', '').lower()
                value = stat.get('value', 0)
                
                if 'cost' in stat_name and value > 0:  # Positive = savings
                    hourly_rate = abs(float(value))
                    return hourly_rate * 730  # Monthly estimate
        
        return 0.0
    
    def _calculate_monthly_cost(self, action: Dict) -> float:
        """Calculate monthly additional cost from action (only for upsizing)."""
        stats = action.get('stats', [])
        for stat in stats:
            if isinstance(stat, dict):
                stat_name = stat.get('name', '').lower()
                value = stat.get('value', 0)
                
                if 'cost' in stat_name and value < 0:  # Negative = additional cost
                    hourly_rate = abs(float(value))
                    return hourly_rate * 730  # Monthly estimate
        
        return 0.0
    
    def _extract_disk_tier(self, tier_string: str) -> str:
        """Extract disk tier from string."""
        if not tier_string or tier_string == 'N/A':
            return 'N/A'
        
        tier_upper = tier_string.upper()
        
        if 'PREMIUM' in tier_upper and 'SSD' in tier_upper:
            return 'Premium SSD'
        elif 'STANDARD' in tier_upper and 'SSD' in tier_upper:
            return 'Standard SSD'
        elif 'STANDARD' in tier_upper and 'HDD' in tier_upper:
            return 'Standard HDD'
        elif 'ULTRA' in tier_upper:
            return 'Ultra SSD'
        
        return tier_string
    
    def _get_attached_vm(self, action: Dict) -> str:
        """Get the VM name that the disk is attached to."""
        target = action.get('target', {})
        
        # Try to get from providers
        providers = target.get('providers', [])
        if providers and isinstance(providers, list):
            for provider in providers:
                if isinstance(provider, dict):
                    vm_name = provider.get('displayName', '')
                    if vm_name:
                        return vm_name
        
        return 'N/A'
    
    def _get_action_link(self, action: Dict) -> str:
        """Generate link to entity page with actions tab in Turbonomic UI."""
        target = action.get('target', {})
        entity_uuid = target.get('uuid', '')
        if entity_uuid:
            # Link to the VM entity page with actions tab selected
            return f"{self.turbo_url}/app/#/view/main/live/{entity_uuid}/overview?selectedTab=actions"
        return 'N/A'
    
    def _sort_actions(self, actions: List[Dict]) -> List[Dict]:
        """Sort actions by environment (Dev>UAT>Pre-Prod>Prod>DR) then by savings (largest to smallest)."""
        return sorted(actions, key=lambda x: (
            self.env_sort_order.get(x.get('Environment', 'Unmapped'), 999),
            -x.get('Monthly Savings', 0)  # Negative for descending order
        ))
    
    def generate_action_plan(self) -> Dict[str, List[Dict]]:
        """Generate categorized action plan."""
        vm_actions = self.get_vm_actions()
        storage_actions = self.get_storage_actions()
        
        # Initialize categories
        must_do_actions = []
        cost_optimization_actions = []
        reliability_investment_actions = []
        
        # Process VM actions
        for action in vm_actions:
            target = action.get('target', {})
            vm_name = target.get('displayName', 'N/A')
            
            # Get customer ID and map to friendly name
            customer_id = self._get_tag_value(target, 'CustomerID')
            customer_name = self._map_customer_name(customer_id)
            
            environment = self._determine_environment(action)
            action_type = self._determine_action_type(action)
            
            current_entity = action.get('currentEntity', {})
            new_entity = action.get('newEntity', {})
            current_config = current_entity.get('displayName', 'N/A')
            recommended_config = new_entity.get('displayName', 'N/A')
            
            monthly_savings = self._calculate_monthly_savings(action) if action_type == 'Downsize' else 0.0
            net_add_cost = self._calculate_monthly_cost(action) if action_type == 'Upsize' else 0.0
            
            action_data = {
                'Type': 'VM Rightsizing',
                'Server/Resource Name': vm_name,
                'Customer ID': customer_id,
                'Customer Friendly Name': customer_name,
                'Environment': environment,
                'Action': action_type,
                'Current Configuration': current_config,
                'Recommended Configuration': recommended_config,
                'Monthly Savings': monthly_savings,
                'Net Add Cost': net_add_cost,
                'Justification': action.get('details', 'N/A'),
                'Risk': action.get('risk', {}).get('severity', 'N/A'),
                'Action Details Link': self._get_action_link(action),
                'UUID': target.get('uuid', 'N/A')
            }
            
            # Categorize based on action type and environment
            if action_type == 'Upsize':
                # All upsizing goes to Reliability Investment (requires budget)
                reliability_investment_actions.append(action_data)
            elif action_type == 'Downsize':
                # Validated downsizing goes to Must-Do if high confidence
                risk_severity = action.get('risk', {}).get('severity', '').upper()
                if risk_severity in ['CRITICAL', 'MAJOR']:
                    must_do_actions.append(action_data)
                else:
                    cost_optimization_actions.append(action_data)
        
        # Process storage actions
        for action in storage_actions:
            target = action.get('target', {})
            disk_name = target.get('displayName', 'N/A')
            vm_name = self._get_attached_vm(action)
            
            # Get customer ID and map to friendly name
            customer_id = self._get_tag_value(target, 'CustomerID')
            customer_name = self._map_customer_name(customer_id)
            
            environment = self._determine_environment(action)
            
            current_entity = action.get('currentEntity', {})
            new_entity = action.get('newEntity', {})
            current_tier = self._extract_disk_tier(current_entity.get('displayName', 'N/A'))
            recommended_tier = self._extract_disk_tier(new_entity.get('displayName', 'N/A'))
            
            # Skip if not a tier change
            if current_tier == recommended_tier or current_tier == 'N/A' or recommended_tier == 'N/A':
                continue
            
            monthly_savings = self._calculate_monthly_savings(action)
            
            action_data = {
                'Type': 'Disk Tier Optimization',
                'Server/Resource Name': f"{vm_name} - {disk_name}",
                'Customer ID': customer_id,
                'Customer Friendly Name': customer_name,
                'Environment': environment,
                'Action': f"{current_tier} → {recommended_tier}",
                'Current Configuration': current_tier,
                'Recommended Configuration': recommended_tier,
                'Monthly Savings': monthly_savings,
                'Net Add Cost': 0.0,
                'Justification': action.get('details', 'N/A'),
                'Risk': action.get('risk', {}).get('severity', 'N/A'),
                'Action Details Link': self._get_action_link(action),
                'UUID': target.get('uuid', 'N/A')
            }
            
            # Categorize
            if environment in ['Dev', 'UAT'] and 'Premium' in current_tier and 'Standard' in recommended_tier:
                # Policy violation - Premium SSD in DEV/UAT
                action_data['Justification'] = 'POLICY VIOLATION: Premium SSD not allowed in DEV/UAT environments'
                must_do_actions.append(action_data)
            elif environment in ['Pre-Prod', 'Prod', 'DR']:
                # Production environments require review
                action_data['Justification'] = f"Review Required: {action.get('details', 'Disk tier optimization')}"
                cost_optimization_actions.append(action_data)
            else:
                cost_optimization_actions.append(action_data)
        
        # Sort all action lists
        must_do_actions = self._sort_actions(must_do_actions)
        cost_optimization_actions = self._sort_actions(cost_optimization_actions)
        reliability_investment_actions = self._sort_actions(reliability_investment_actions)
        
        return {
            'must_do': must_do_actions,
            'cost_optimization': cost_optimization_actions,
            'reliability_investment': reliability_investment_actions
        }
    
    def export_to_excel(self, action_plan: Dict[str, List[Dict]], output_file: str):
        """Export action plan to formatted Excel file."""
        if not PANDAS_AVAILABLE:
            print("Error: pandas and openpyxl required for Excel export")
            return
        
        # Create Excel writer
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            # Summary sheet
            self._create_summary_sheet(writer, action_plan)
            
            # Must-Do Actions sheet
            if action_plan['must_do']:
                df_must_do = pd.DataFrame(action_plan['must_do'])
                df_must_do.to_excel(writer, sheet_name='1. Must-Do Actions', index=False)
                self._format_sheet(writer.book['1. Must-Do Actions'], 'must_do')
            
            # Cost Optimization sheet
            if action_plan['cost_optimization']:
                df_cost = pd.DataFrame(action_plan['cost_optimization'])
                df_cost.to_excel(writer, sheet_name='2. Cost Optimization', index=False)
                self._format_sheet(writer.book['2. Cost Optimization'], 'cost_optimization')
            
            # Reliability Investment sheet
            if action_plan['reliability_investment']:
                df_reliability = pd.DataFrame(action_plan['reliability_investment'])
                df_reliability.to_excel(writer, sheet_name='3. Reliability Investment', index=False)
                self._format_sheet(writer.book['3. Reliability Investment'], 'reliability')
        
        print(f"\n✓ Monthly Action Plan exported to: {output_file}")
    
    def _create_summary_sheet(self, writer, action_plan: Dict[str, List[Dict]]):
        """Create summary sheet with statistics."""
        must_do_count = len(action_plan['must_do'])
        cost_opt_count = len(action_plan['cost_optimization'])
        reliability_count = len(action_plan['reliability_investment'])
        
        must_do_savings = sum(a['Monthly Savings'] for a in action_plan['must_do'])
        cost_opt_savings = sum(a['Monthly Savings'] for a in action_plan['cost_optimization'])
        reliability_cost = sum(a['Net Add Cost'] for a in action_plan['reliability_investment'])
        
        summary_data = {
            'Category': [
                'Must-Do Actions',
                'Cost Optimization',
                'Reliability Investment',
                '',
                'Total Actions',
                'Total Monthly Savings',
                'Total Investment Required',
                'Net Monthly Impact'
            ],
            'Count': [
                must_do_count,
                cost_opt_count,
                reliability_count,
                '',
                must_do_count + cost_opt_count + reliability_count,
                '',
                '',
                ''
            ],
            'Monthly Impact ($)': [
                f"${must_do_savings:,.2f}",
                f"${cost_opt_savings:,.2f}",
                f"-${reliability_cost:,.2f}",
                '',
                '',
                f"${must_do_savings + cost_opt_savings:,.2f}",
                f"${reliability_cost:,.2f}",
                f"${must_do_savings + cost_opt_savings - reliability_cost:,.2f}"
            ],
            'Description': [
                'Policy violations + validated downsizing',
                'Recommended downsizing across all environments',
                'Prod upsizing with budget impact',
                '',
                '',
                'Total potential savings',
                'Total additional cost for reliability',
                'Net financial impact'
            ]
        }
        
        df_summary = pd.DataFrame(summary_data)
        df_summary.to_excel(writer, sheet_name='Summary', index=False)
        
        # Format summary sheet
        ws = writer.book['Summary']
        self._format_summary_sheet(ws)
    
    def _format_summary_sheet(self, ws):
        """Format the summary sheet."""
        # Header formatting
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Category formatting
        category_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        category_font = Font(bold=True, size=11)
        
        for row in [2, 3, 4]:  # Must-Do, Cost Opt, Reliability rows
            ws[f'A{row}'].fill = category_fill
            ws[f'A{row}'].font = category_font
        
        # Total rows formatting
        total_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        total_font = Font(bold=True, size=11)
        
        for row in [6, 7, 8, 9]:  # Total rows
            for col in ['A', 'B', 'C', 'D']:
                ws[f'{col}{row}'].fill = total_fill
                ws[f'{col}{row}'].font = total_font
        
        # Column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 50
        
        # Borders
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in ws.iter_rows(min_row=1, max_row=9, min_col=1, max_col=4):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(vertical='center', wrap_text=True)
    
    def _format_sheet(self, ws, sheet_type: str):
        """Format action sheets."""
        # Header formatting
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Column widths
        column_widths = {
            'A': 25,  # Type
            'B': 35,  # Server/Resource Name
            'C': 20,  # Customer ID
            'D': 30,  # Customer Friendly Name
            'E': 15,  # Environment
            'F': 20,  # Action
            'G': 25,  # Current Configuration
            'H': 25,  # Recommended Configuration
            'I': 18,  # Monthly Savings
            'J': 18,  # Net Add Cost
            'K': 50,  # Justification
            'L': 12,  # Risk
            'M': 50,  # Action Details Link
            'N': 38   # UUID
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # Borders and alignment
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=14):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(vertical='center', wrap_text=True)
        
        # Make Action Details Link column clickable (Column M)
        for row in range(2, ws.max_row + 1):
            link_cell = ws[f'M{row}']
            if link_cell.value and link_cell.value != 'N/A':
                link_cell.hyperlink = link_cell.value
                link_cell.style = 'Hyperlink'
        
        # Highlight policy violations in Must-Do sheet (red background)
        if sheet_type == 'must_do':
            violation_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            for row in range(2, ws.max_row + 1):
                justification = ws[f'K{row}'].value
                if justification and 'POLICY VIOLATION' in str(justification):
                    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N']:
                        ws[f'{col}{row}'].fill = violation_fill
        
        # Freeze panes
        ws.freeze_panes = 'A2'


def main():
    parser = argparse.ArgumentParser(
        description='Generate Monthly Action Plan from Turbonomic',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Generate monthly action plan
  python3 generate_monthly_action_plan.py --url https://turbo.example.com --jsessionid SESSION_ID
  
  # With customer mapping
  python3 generate_monthly_action_plan.py --url https://turbo.example.com --jsessionid SESSION_ID --customer-mapping customer_mapping.json
  
  # Custom output file
  python3 generate_monthly_action_plan.py --url https://turbo.example.com --jsessionid SESSION_ID --output monthly_plan_jan2026.xlsx
        """
    )
    
    parser.add_argument('--url', required=True, help='Turbonomic instance URL')
    parser.add_argument('--jsessionid', required=True, help='Session ID from Turbonomic login')
    parser.add_argument('--output', default=None, help='Output filename (default: Monthly_Action_Plan_YYYYMMDD_HHMMSS.xlsx)')
    parser.add_argument('--customer-mapping', default='customer_mapping.json', help='Customer mapping JSON file (default: customer_mapping.json)')
    
    args = parser.parse_args()
    
    # Generate default output filename if not provided
    if not args.output:
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        args.output = f"Monthly_Action_Plan_{timestamp}.xlsx"
    
    print("=" * 80)
    print("TURBONOMIC MONTHLY ACTION PLAN GENERATOR")
    print("=" * 80)
    print(f"Turbonomic URL: {args.url}")
    print(f"Output file: {args.output}")
    print("=" * 80)
    print()
    
    # Use customer mapping if file exists
    customer_mapping_file = args.customer_mapping if os.path.exists(args.customer_mapping) else None
    
    # Create generator
    generator = MonthlyActionPlanGenerator(
        turbo_url=args.url,
        jsessionid=args.jsessionid,
        customer_mapping_file=customer_mapping_file
    )
    
    # Generate action plan
    print("Generating monthly action plan...")
    action_plan = generator.generate_action_plan()
    
    # Print summary
    print("\n" + "=" * 80)
    print("ACTION PLAN SUMMARY")
    print("=" * 80)
    print(f"Must-Do Actions:           {len(action_plan['must_do'])}")
    print(f"Cost Optimization Actions: {len(action_plan['cost_optimization'])}")
    print(f"Reliability Investments:   {len(action_plan['reliability_investment'])}")
    print(f"Total Actions:             {len(action_plan['must_do']) + len(action_plan['cost_optimization']) + len(action_plan['reliability_investment'])}")
    
    must_do_savings = sum(a['Monthly Savings'] for a in action_plan['must_do'])
    cost_opt_savings = sum(a['Monthly Savings'] for a in action_plan['cost_optimization'])
    reliability_cost = sum(a['Net Add Cost'] for a in action_plan['reliability_investment'])
    
    print(f"\nMust-Do Savings:           ${must_do_savings:,.2f}/month")
    print(f"Cost Optimization Savings: ${cost_opt_savings:,.2f}/month")
    print(f"Reliability Investment:    ${reliability_cost:,.2f}/month")
    print(f"Net Monthly Impact:        ${must_do_savings + cost_opt_savings - reliability_cost:,.2f}/month")
    print("=" * 80)
    
    # Export to Excel
    generator.export_to_excel(action_plan, args.output)
    
    print("\n✓ Monthly Action Plan generation complete!")
    print(f"\nNext steps:")
    print(f"1. Review 'Must-Do Actions' sheet for policy violations and validated downsizing")
    print(f"2. Evaluate 'Cost Optimization' sheet for additional savings opportunities")
    print(f"3. Assess 'Reliability Investment' sheet for production upsizing with budget approval")


if __name__ == '__main__':
    main()

# Made with Bob
