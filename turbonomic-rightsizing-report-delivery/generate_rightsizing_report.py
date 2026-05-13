#!/usr/bin/env python3
"""
Turbonomic Rightsizing Recommendations Report Generator v4 (Consolidated Excel Format)

This script generates consolidated rightsizing reports from Turbonomic with:
- Single Excel workbook with multiple sheets (Summary, Comprehensive, Dev, UAT, Pre-Prod, Prod, DR, Unmapped)
- Professional formatting matching Monthly Action Plan style
- Clickable action links with corrected URL format
- Azure-only filtering
- BusinessAccount extraction
- Environment parsing from BusinessAccount name

Requirements:
    pip install requests pandas openpyxl

Usage:
    python3 generate_rightsizing_report_v4.py --url https://your-turbo-instance.com --jsessionid <SESSION_ID> --output-dir reports/
"""

import requests
import json
import argparse
import sys
import re
import os
from datetime import datetime
from typing import List, Dict, Optional, Tuple
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

try:
    import pandas as pd
    from openpyxl import Workbook
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False


class TurbonomicRightsizingReport:
    """Generate consolidated rightsizing recommendations report from Turbonomic API."""
    
    def __init__(self, turbo_url: str, jsessionid: str, customer_mapping_file: Optional[str] = None):
        self.turbo_url = turbo_url.rstrip('/')
        self.session = requests.Session()
        self.session.cookies.set('JSESSIONID', jsessionid)
        self.session.headers.update({
            'Content-Type': 'application/json',
            'Accept': 'application/json'
        })
        self.customer_mapping = self._load_customer_mapping(customer_mapping_file)
        
        # Environment sort order for consistent ordering
        self.env_sort_order = {'Dev': 1, 'UAT': 2, 'Pre-Prod': 3, 'Prod': 4, 'DR': 5, 'Unmapped': 6}
    
    def _load_customer_mapping(self, mapping_file: Optional[str]) -> Dict[str, str]:
        """Load customer ID to name mapping from JSON file."""
        if not mapping_file:
            return {}
        
        try:
            with open(mapping_file, 'r', encoding='utf-8') as f:
                mapping = json.load(f)
                print(f"✓ Loaded {len(mapping)} customer mappings from {mapping_file}")
                return mapping
        except FileNotFoundError:
            print(f"Warning: Customer mapping file not found: {mapping_file}")
            return {}
        except json.JSONDecodeError as e:
            print(f"Warning: Invalid JSON in customer mapping file: {e}")
            return {}
        except Exception as e:
            print(f"Warning: Error loading customer mapping file: {e}")
            return {}
    
    def _map_customer_name(self, customer_id: str) -> str:
        """Map customer ID to friendly name using the mapping file."""
        if not customer_id or customer_id == 'N/A':
            return 'N/A'
        
        # Try exact match first
        if customer_id in self.customer_mapping:
            return self.customer_mapping[customer_id]
        
        # Return original ID if no mapping found
        return customer_id
    
    def get_recommended_actions(self, fetch_all: bool = True) -> List[Dict]:
        """Fetch all recommended actions with pagination support."""
        url = f"{self.turbo_url}/api/v3/markets/Market/actions"
        
        payload = {
            "actionStateList": ["READY", "QUEUED", "IN_PROGRESS", "ACCEPTED"],
            "actionTypeList": ["RESIZE", "SCALE"],
            "relatedEntityTypes": ["VirtualMachine"],
            "environmentType": "CLOUD",
            "detailLevel": "EXECUTION"
        }
        
        all_actions = []
        cursor = 0
        page_size = 500
        
        try:
            print(f"Fetching recommended actions from {url}...")
            
            while True:
                params = {
                    "ascending": "false",
                    "cursor": str(cursor),
                    "disable_hateoas": "true",
                    "forceExpansionOfAggregatedEntities": "true",
                    "order_by": "savings",
                    "limit": str(page_size)
                }
                
                response = self.session.post(url, json=payload, params=params)
                response.raise_for_status()
                
                data = response.json()
                actions = data if isinstance(data, list) else []
                
                if not actions:
                    break
                
                all_actions.extend(actions)
                print(f"  Retrieved {len(actions)} actions (total: {len(all_actions)})")
                
                if len(actions) < page_size or not fetch_all:
                    break
                
                cursor += page_size
            
            print(f"Total actions retrieved: {len(all_actions)}")
            return all_actions
            
        except requests.exceptions.RequestException as e:
            print(f"Error fetching actions: {e}")
            if hasattr(e, 'response') and e.response is not None:
                print(f"Response: {e.response.text}")
            return []
    
    def _get_cloud_provider(self, action: Dict) -> str:
        """Determine cloud provider from action data."""
        target = action.get('target', {})
        discovered_by = target.get('discoveredBy', {})
        probe_type = discovered_by.get('type', '').lower()
        
        if 'azure' in probe_type:
            return 'Azure'
        elif 'aws' in probe_type:
            return 'AWS'
        elif 'gcp' in probe_type or 'google' in probe_type:
            return 'GCP'
        
        # Check aspects
        aspects = target.get('aspects', {})
        if 'cloudAspect' in aspects:
            return 'Azure'  # Based on your data structure
        
        return 'Unknown'
    
    def _get_business_account(self, action: Dict) -> str:
        """Extract BusinessAccount name from cloudAspect."""
        target = action.get('target', {})
        aspects = target.get('aspects', {})
        cloud_aspect = aspects.get('cloudAspect', {})
        business_account = cloud_aspect.get('businessAccount', {})
        
        return business_account.get('displayName', 'N/A')
    
    def _parse_environment_from_account(self, account_name: str) -> Optional[str]:
        """
        Parse environment from BusinessAccount name.
        
        Common patterns:
        - Contains 'PRD', 'PROD', 'Production' -> Prod
        - Contains 'DEV', 'Development' -> Dev
        - Contains 'UAT', 'User Acceptance' -> UAT
        - Contains 'PRE-PROD', 'PREPROD', 'Pre-Production' -> Pre-Prod
        - Contains 'DR', 'Disaster Recovery' -> DR
        
        Returns None if no match found.
        """
        account_upper = account_name.upper()
        
        # Check for Production (most specific first)
        if any(x in account_upper for x in ['PRODUCTION', ' PRD ', '-PRD-', '_PRD_', 'PRD-', '-PRD']):
            return 'Prod'
        
        # Check for Pre-Prod
        if any(x in account_upper for x in ['PRE-PROD', 'PREPROD', 'PRE-PRODUCTION', 'PREPRODUCTION']):
            return 'Pre-Prod'
        
        # Check for DR
        if any(x in account_upper for x in [' DR ', '-DR-', '_DR_', 'DR-', '-DR', 'DISASTER RECOVERY', 'DISASTERRECOVERY']):
            return 'DR'
        
        # Check for UAT
        if any(x in account_upper for x in ['UAT', 'USER ACCEPTANCE', 'USERACCEPTANCE']):
            return 'UAT'
        
        # Check for Dev
        if any(x in account_upper for x in ['DEVELOPMENT', ' DEV ', '-DEV-', '_DEV_', 'DEV-', '-DEV']):
            return 'Dev'
        
        return None  # No match found
    
    def _determine_environment(self, action: Dict) -> str:
        """
        Determine environment with fallback logic:
        1. Parse from BusinessAccount name
        2. Fall back to environment tag
        3. Mark as 'Unmapped' if neither works
        """
        # Try BusinessAccount first
        account_name = self._get_business_account(action)
        if account_name != 'N/A':
            env_from_account = self._parse_environment_from_account(account_name)
            if env_from_account:
                return env_from_account
        
        # Fall back to environment tag
        target = action.get('target', {})
        env_tag = self._get_tag_value(target, 'environment')
        
        if env_tag and env_tag != 'N/A':
            env_upper = env_tag.upper()
            if 'PROD' in env_upper and 'PRE' not in env_upper:
                return 'Prod'
            elif 'PRE-PROD' in env_upper or 'PREPROD' in env_upper:
                return 'Pre-Prod'
            elif 'UAT' in env_upper:
                return 'UAT'
            elif 'DEV' in env_upper:
                return 'Dev'
            elif 'DR' in env_upper:
                return 'DR'
            # If tag exists but doesn't match patterns, categorize as Unmapped
        
        return 'Unmapped'
    
    def _get_tag_value(self, entity: Dict, tag_key: str) -> str:
        """Extract tag value from entity."""
        tags = entity.get('tags', {})
        
        if isinstance(tags, dict):
            if tag_key in tags:
                tag_value = tags[tag_key]
                if isinstance(tag_value, list) and tag_value:
                    return tag_value[0] if tag_value[0] else 'N/A'
                return str(tag_value) if tag_value else 'N/A'
            
            for key, value in tags.items():
                if key.lower() == tag_key.lower():
                    if isinstance(value, list) and value:
                        return value[0] if value[0] else 'N/A'
                    return str(value) if value else 'N/A'
        
        return 'N/A'
    
    def _determine_action_type(self, action: Dict) -> str:
        """Determine if action is upsize or downsize based on cost impact."""
        # First check cost stats - most reliable indicator
        # In Turbonomic API:
        # Positive cost value = SAVINGS (downsize - reducing resources saves money)
        # Negative cost value = ADDITIONAL COST (upsize - adding resources costs money)
        stats = action.get('stats', [])
        for stat in stats:
            if isinstance(stat, dict):
                stat_name = stat.get('name', '').lower()
                value = stat.get('value', 0)
                
                if 'cost' in stat_name and value != 0:
                    if value > 0:  # Positive = savings = Downsize
                        return 'Downsize'
                    elif value < 0:  # Negative = additional cost = Upsize
                        return 'Upsize'
        
        # Fallback to risk category
        risk = action.get('risk', {})
        risk_sub_category = risk.get('subCategory', '').lower()
        details = action.get('details', '').lower()
        
        if 'underutilized' in risk_sub_category or 'underutilized' in details:
            return 'Downsize'
        
        if 'overutilized' in risk_sub_category or 'overutilized' in details:
            return 'Upsize'
        
        # Try to parse from VM size names
        current_entity = action.get('currentEntity', {})
        new_entity = action.get('newEntity', {})
        current_name = current_entity.get('displayName', '')
        new_name = new_entity.get('displayName', '')
        
        if current_name and new_name:
            current_nums = re.findall(r'\d+', current_name)
            new_nums = re.findall(r'\d+', new_name)
            
            if current_nums and new_nums:
                try:
                    current_size = int(current_nums[0])
                    new_size = int(new_nums[0])
                    if new_size > current_size:
                        return 'Upsize'
                    elif new_size < current_size:
                        return 'Downsize'
                except (ValueError, IndexError):
                    pass
        
        return 'Resize'
    
    def _format_configuration(self, action: Dict) -> str:
        """Format current VM configuration."""
        current_entity = action.get('currentEntity', {})
        return current_entity.get('displayName', 'N/A')
    
    def _format_recommendation(self, action: Dict) -> str:
        """Format recommended configuration."""
        new_entity = action.get('newEntity', {})
        return new_entity.get('displayName', 'N/A')
    
    def _calculate_monthly_savings(self, action: Dict) -> float:
        """Calculate monthly savings (only for downsizing)."""
        stats = action.get('stats', [])
        for stat in stats:
            if isinstance(stat, dict):
                stat_name = stat.get('name', '').lower()
                value = stat.get('value', 0)
                
                if 'cost' in stat_name and value > 0:  # Positive = savings
                    hourly_rate = abs(float(value))
                    return hourly_rate * 730  # Monthly estimate
        
        return 0.0
    
    def _calculate_monthly_cost(self, action: Dict) -> float:
        """Calculate monthly additional cost (only for upsizing)."""
        stats = action.get('stats', [])
        for stat in stats:
            if isinstance(stat, dict):
                stat_name = stat.get('name', '').lower()
                value = stat.get('value', 0)
                
                if 'cost' in stat_name and value < 0:  # Negative = additional cost
                    hourly_rate = abs(float(value))
                    return hourly_rate * 730  # Monthly estimate
        
        return 0.0
    
    def _get_action_link(self, action: Dict) -> str:
        """Generate link to entity page with actions tab in Turbonomic UI."""
        target = action.get('target', {})
        entity_uuid = target.get('uuid', '')
        if entity_uuid:
            # Link to the VM entity page with actions tab selected
            return f"{self.turbo_url}/app/#/view/main/live/{entity_uuid}/overview?selectedTab=actions"
        return 'N/A'
    
    def generate_report_data(self, azure_only: bool = True, 
                            action_type_filter: Optional[str] = None) -> List[Dict]:
        """Generate report data from Turbonomic actions."""
        actions = self.get_recommended_actions()
        
        report_data = []
        
        for action in actions:
            target = action.get('target', {})
            
            # Filter for Azure only
            cloud_provider = self._get_cloud_provider(action)
            if azure_only and cloud_provider != 'Azure':
                continue
            
            # Determine environment
            environment = self._determine_environment(action)
            
            # Determine action type
            action_type = self._determine_action_type(action)
            
            # Apply action type filter
            if action_type_filter:
                if action_type_filter.lower() == 'downsize' and action_type != 'Downsize':
                    continue
                if action_type_filter.lower() == 'upsize' and action_type != 'Upsize':
                    continue
            
            # Get customer ID and map to friendly name
            customer_id = self._get_tag_value(target, 'CustomerID')
            customer_name = self._map_customer_name(customer_id)
            
            # Calculate costs
            monthly_savings = self._calculate_monthly_savings(action) if action_type == 'Downsize' else 0.0
            net_add_cost = self._calculate_monthly_cost(action) if action_type == 'Upsize' else 0.0
            
            # Build report row
            row = {
                'Server Name': target.get('displayName', 'N/A'),
                'Customer ID': customer_id,
                'Customer Friendly Name': customer_name,
                'Current Configuration': self._format_configuration(action),
                'Recommendation': self._format_recommendation(action),
                'Action Type': action_type,
                'Monthly Savings': monthly_savings,
                'Net Add Cost': net_add_cost,
                'Environment': environment,
                'Action Details Link': self._get_action_link(action),
                'Business Account': self._get_business_account(action),
                'Cloud Provider': cloud_provider,
                'Action State': action.get('actionState', 'N/A'),
                'Risk': action.get('risk', {}).get('severity', 'N/A'),
                'Details': action.get('details', 'N/A'),
                'UUID': target.get('uuid', 'N/A')
            }
            
            report_data.append(row)
        
        return report_data
    
    def _sort_data(self, data: List[Dict]) -> List[Dict]:
        """Sort data by environment then by savings (largest to smallest)."""
        return sorted(data, key=lambda x: (
            self.env_sort_order.get(x.get('Environment', 'Unmapped'), 999),
            -x.get('Monthly Savings', 0)  # Negative for descending order
        ))
    
    def export_consolidated_excel(self, all_data: List[Dict], filename: str):
        """Export consolidated report to Excel with multiple sheets and professional formatting."""
        if not PANDAS_AVAILABLE:
            print("Error: pandas and openpyxl required for Excel export")
            return
        
        if not all_data:
            print(f"No data to export to {filename}")
            return
        
        # Sort all data
        all_data = self._sort_data(all_data)
        
        # Create Excel writer
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            # Create Summary sheet
            self._create_summary_sheet(writer, all_data)
            
            # Create Comprehensive sheet (all data)
            df_comprehensive = pd.DataFrame(all_data)
            df_comprehensive.to_excel(writer, sheet_name='Comprehensive', index=False)
            self._format_data_sheet(writer.book['Comprehensive'])
            
            # Create environment-specific sheets
            environments = ['Dev', 'UAT', 'Pre-Prod', 'Prod', 'DR', 'Unmapped']
            for env in environments:
                env_data = [row for row in all_data if row['Environment'] == env]
                if env_data:
                    df_env = pd.DataFrame(env_data)
                    df_env.to_excel(writer, sheet_name=env, index=False)
                    self._format_data_sheet(writer.book[env])
        
        print(f"\n✓ Consolidated report exported to: {filename}")
        print(f"  Sheets: Summary, Comprehensive, {', '.join(environments)}")
    
    def _create_summary_sheet(self, writer, all_data: List[Dict]):
        """Create summary sheet with statistics."""
        # Calculate overall statistics
        total_actions = len(all_data)
        downsizes = len([r for r in all_data if r['Action Type'] == 'Downsize'])
        upsizes = len([r for r in all_data if r['Action Type'] == 'Upsize'])
        other = total_actions - downsizes - upsizes
        
        total_monthly_savings = sum(r['Monthly Savings'] for r in all_data)
        total_add_cost = sum(r['Net Add Cost'] for r in all_data)
        net_impact = total_monthly_savings - total_add_cost
        
        # Environment breakdown
        env_breakdown = {}
        env_savings = {}
        env_costs = {}
        env_downsizes = {}
        env_upsizes = {}
        
        for row in all_data:
            env = row.get('Environment', 'Unknown')
            env_breakdown[env] = env_breakdown.get(env, 0) + 1
            env_savings[env] = env_savings.get(env, 0) + row['Monthly Savings']
            env_costs[env] = env_costs.get(env, 0) + row['Net Add Cost']
            
            if row['Action Type'] == 'Downsize':
                env_downsizes[env] = env_downsizes.get(env, 0) + 1
            elif row['Action Type'] == 'Upsize':
                env_upsizes[env] = env_upsizes.get(env, 0) + 1
        
        # Build summary data
        summary_rows = []
        
        # Overall summary
        summary_rows.append({
            'Category': 'OVERALL SUMMARY',
            'Count': '',
            'Downsizes': '',
            'Upsizes': '',
            'Monthly Savings': '',
            'Monthly Add Cost': '',
            'Net Impact': ''
        })
        summary_rows.append({
            'Category': 'Total Recommendations',
            'Count': total_actions,
            'Downsizes': downsizes,
            'Upsizes': upsizes,
            'Monthly Savings': f"${total_monthly_savings:,.2f}",
            'Monthly Add Cost': f"${total_add_cost:,.2f}",
            'Net Impact': f"${net_impact:,.2f}"
        })
        summary_rows.append({
            'Category': '',
            'Count': '',
            'Downsizes': '',
            'Upsizes': '',
            'Monthly Savings': '',
            'Monthly Add Cost': '',
            'Net Impact': ''
        })
        
        # Environment breakdown
        summary_rows.append({
            'Category': 'ENVIRONMENT BREAKDOWN',
            'Count': '',
            'Downsizes': '',
            'Upsizes': '',
            'Monthly Savings': '',
            'Monthly Add Cost': '',
            'Net Impact': ''
        })
        
        for env in sorted(env_breakdown.keys(), key=lambda x: self.env_sort_order.get(x, 999)):
            count = env_breakdown[env]
            savings = env_savings.get(env, 0)
            costs = env_costs.get(env, 0)
            down = env_downsizes.get(env, 0)
            up = env_upsizes.get(env, 0)
            net = savings - costs
            
            summary_rows.append({
                'Category': env,
                'Count': count,
                'Downsizes': down,
                'Upsizes': up,
                'Monthly Savings': f"${savings:,.2f}",
                'Monthly Add Cost': f"${costs:,.2f}",
                'Net Impact': f"${net:,.2f}"
            })
        
        df_summary = pd.DataFrame(summary_rows)
        df_summary.to_excel(writer, sheet_name='Summary', index=False)
        
        # Format summary sheet
        ws = writer.book['Summary']
        self._format_summary_sheet(ws)
    
    def _format_summary_sheet(self, ws):
        """Format the summary sheet."""
        # Header formatting
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Section header formatting (OVERALL SUMMARY, ENVIRONMENT BREAKDOWN)
        section_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        section_font = Font(bold=True, size=11)
        
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            if row[0].value in ['OVERALL SUMMARY', 'ENVIRONMENT BREAKDOWN']:
                for cell in row:
                    cell.fill = section_fill
                    cell.font = section_font
        
        # Environment row formatting
        env_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            if row[0].value in ['Dev', 'UAT', 'Pre-Prod', 'Prod', 'DR', 'Unmapped']:
                for cell in row:
                    cell.fill = env_fill
        
        # Column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 20
        ws.column_dimensions['G'].width = 20
        
        # Borders
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=7):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(vertical='center', wrap_text=True)
        
        # Freeze panes
        ws.freeze_panes = 'A2'
    
    def _format_data_sheet(self, ws):
        """Format data sheets (Comprehensive and environment-specific)."""
        # Header formatting
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Column widths
        column_widths = {
            'A': 35,  # Server Name
            'B': 20,  # Customer ID
            'C': 30,  # Customer Friendly Name
            'D': 25,  # Current Configuration
            'E': 25,  # Recommendation
            'F': 15,  # Action Type
            'G': 18,  # Monthly Savings
            'H': 18,  # Net Add Cost
            'I': 15,  # Environment
            'J': 50,  # Action Details Link
            'K': 30,  # Business Account
            'L': 15,  # Cloud Provider
            'M': 15,  # Action State
            'N': 12,  # Risk
            'O': 50,  # Details
            'P': 38   # UUID
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # Borders and alignment
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=16):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(vertical='center', wrap_text=True)
        
        # Make Action Details Link column clickable (column J)
        for row in range(2, ws.max_row + 1):
            link_cell = ws[f'J{row}']
            if link_cell.value and link_cell.value != 'N/A':
                link_cell.hyperlink = link_cell.value
                link_cell.style = 'Hyperlink'
        
        # Freeze panes
        ws.freeze_panes = 'A2'
    
    def print_summary(self, report_data: List[Dict], title: str = "RIGHTSIZING REPORT SUMMARY"):
        """Print report summary statistics."""
        if not report_data:
            print(f"No data in {title}")
            return
        
        total_actions = len(report_data)
        downsizes = len([r for r in report_data if r['Action Type'] == 'Downsize'])
        upsizes = len([r for r in report_data if r['Action Type'] == 'Upsize'])
        
        total_monthly_savings = sum(r['Monthly Savings'] for r in report_data)
        total_add_cost = sum(r['Net Add Cost'] for r in report_data)
        
        print("\n" + "="*80)
        print(title)
        print("="*80)
        print(f"Total Recommendations: {total_actions}")
        print(f"  - Downsizing: {downsizes}")
        print(f"  - Upsizing: {upsizes}")
        print(f"  - Other: {total_actions - downsizes - upsizes}")
        print(f"\nEstimated Monthly Savings: ${total_monthly_savings:,.2f}")
        print(f"Estimated Monthly Add Cost: ${total_add_cost:,.2f}")
        print(f"Net Monthly Impact: ${total_monthly_savings - total_add_cost:,.2f}")
        
        environments = {}
        for row in report_data:
            env = row.get('Environment', 'Unknown')
            environments[env] = environments.get(env, 0) + 1
        
        if environments:
            print(f"\nEnvironment Breakdown:")
            for env in sorted(environments.keys(), key=lambda x: self.env_sort_order.get(x, 999)):
                print(f"  - {env}: {environments[env]}")
        
        print("="*80)


def main():
    parser = argparse.ArgumentParser(
        description='Generate Turbonomic rightsizing reports v4 (Consolidated Excel Format)',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Generate consolidated report
  python3 generate_rightsizing_report_v4.py --url https://turbo.example.com --jsessionid <ID> --output-dir reports/
  
  # With customer mapping
  python3 generate_rightsizing_report_v4.py --url https://turbo.example.com --jsessionid <ID> --customer-mapping customer_mapping.json
  
  # Custom output filename
  python3 generate_rightsizing_report_v4.py --url https://turbo.example.com --jsessionid <ID> --output rightsizing_jan2026.xlsx
        """
    )
    
    parser.add_argument('--url', required=True, help='Turbonomic instance URL')
    parser.add_argument('--jsessionid', required=True, help='JSESSIONID from login')
    parser.add_argument('--output-dir', default='.', help='Output directory for reports (default: current directory)')
    parser.add_argument('--output', help='Custom output filename (default: Rightsizing_Report_TIMESTAMP.xlsx)')
    parser.add_argument('--action-type', choices=['upsize', 'downsize'], help='Filter by action type')
    parser.add_argument('--all-clouds', action='store_true', help='Include all cloud providers (not just Azure)')
    parser.add_argument('--customer-mapping', default='customer_mapping.json', help='Path to customer ID to name mapping JSON file (default: customer_mapping.json)')
    
    args = parser.parse_args()
    
    try:
        # Validate and create output directory
        output_dir = args.output_dir
        
        # Check if output_dir is actually a file (common mistake)
        if os.path.isfile(output_dir):
            print(f"Error: '{output_dir}' is a file, not a directory.")
            print(f"Use --output-dir to specify a directory, not a filename.")
            print(f"Example: --output-dir ./reports/")
            return 1
        
        # Create output directory
        os.makedirs(output_dir, exist_ok=True)
        
        # Use customer mapping if file exists
        customer_mapping_file = args.customer_mapping if os.path.exists(args.customer_mapping) else None
        
        report = TurbonomicRightsizingReport(args.url, args.jsessionid, customer_mapping_file)
        
        print("Generating consolidated rightsizing report...")
        all_data = report.generate_report_data(
            azure_only=not args.all_clouds,
            action_type_filter=args.action_type
        )
        
        if not all_data:
            print("No data found. Exiting.")
            return 1
        
        # Generate consolidated report
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        if args.output:
            output_file = os.path.join(output_dir, args.output)
        else:
            output_file = os.path.join(output_dir, f"Rightsizing_Report_{timestamp}.xlsx")
        
        report.export_consolidated_excel(all_data, output_file)
        report.print_summary(all_data, "CONSOLIDATED RIGHTSIZING REPORT SUMMARY")
        
        print(f"\n{'='*80}")
        print("Report generation complete!")
        print(f"{'='*80}")
        
        return 0
        
    except KeyboardInterrupt:
        print("\n\nOperation cancelled by user")
        return 1
    except Exception as e:
        print(f"\nError: {e}")
        import traceback
        traceback.print_exc()
        return 1


if __name__ == '__main__':
    sys.exit(main())

# Made with Bob
